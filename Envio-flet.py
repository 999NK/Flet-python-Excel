import flet as ft
from flet import Page, TextField, ElevatedButton, Column, DatePicker, icons
import openpyxl
from openpyxl import Workbook
import os
import datetime

def create_or_open_excel(file_name):
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Idade", "Data e Hora", "Opções"])  # Cabeçalhos
        wb.save(file_name)
    return wb

def main(page: Page):
    page.theme_mode = ft.ThemeMode.DARK
    page.window_width = 600  # Define a largura da janela
    page.window_height = 800  # Define a altura da janela
    file_name = "data.xlsx"
    wb = create_or_open_excel(file_name)
    
    table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("Nome")),
            ft.DataColumn(ft.Text("Idade")),
            ft.DataColumn(ft.Text("Data e Hora")),
            ft.DataColumn(ft.Text("Opções")),
            ft.DataColumn(ft.Text("Ações")),
        ],
        rows=[],
    )

    def add_to_table(row):
        delete_button = ft.IconButton(
            icon=icons.DELETE,
            on_click=lambda e: delete_row(e, row)
        )
        table.rows.append(ft.DataRow(cells=[ft.DataCell(ft.Text(cell)) for cell in row] + [ft.DataCell(delete_button)]))
        page.update()

    def delete_row(e, row):
        # Remove a linha da tabela
        table.rows = [r for r in table.rows if r.cells[0].content.value != row[0]]
        # Remove a linha do Excel
        ws = wb.active
        for r in ws.iter_rows(min_row=2):  # Pular a linha de cabeçalho
            if r[0].value == row[0] and r[1].value == row[1] and r[2].value == row[2] and r[3].value == row[3]:
                ws.delete_rows(r[0].row, 1)
                wb.save(file_name)
                break
        page.update()

    def add_to_excel(e):
        selected_date = date_picker.value
        formatted_date = selected_date.strftime("%d/%m/%Y")
        ws = wb.active
        ws.append([name_field.value, age_field.value, formatted_date, dropdown.value])
        wb.save(file_name)
        # Adiciona o novo item à tabela
        add_to_table([name_field.value, age_field.value, formatted_date, dropdown.value])
        # Limpa os campos de entrada
        name_field.value = ""
        age_field.value = ""
        date_picker.value = datetime.datetime.now()
        dropdown.value = None
        page.update()

    name_field = TextField(label="Nome")
    age_field = TextField(label="Idade")
    date_picker = DatePicker(value=datetime.datetime.now())

    def open_date_picker(e):
        date_picker.pick_date()
        
    date_button = ElevatedButton(
        "Selecionar Data",
        icon=icons.CALENDAR_MONTH,
        on_click=open_date_picker
    )

    dropdown = ft.Dropdown(
        label="Cor",
        hint_text="Escolha sua cor favorita",
        options=[
            ft.dropdown.Option("Vermelho"),
            ft.dropdown.Option("Verde"),
            ft.dropdown.Option("Azul"),
        ]
    )

    add_button = ElevatedButton(text="Adicionar ao Excel", on_click=add_to_excel)
    buttons_row = ft.Row(
        controls=[
            add_button,
            date_button
        ],
        alignment="center",
        spacing=10  # Ajuste o espaçamento entre os botões conforme necessário
    )
    page.add(
        Column(
            controls=[
                name_field,
                age_field,
                dropdown,
                buttons_row,
                date_picker,
                table  # Adiciona a tabela à coluna
            ],
            spacing=20,
            horizontal_alignment="center"
        )
    )
ft.app(target=main)
#ft.app(target=main, view=ft.WEB_BROWSER)
